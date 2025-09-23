from flask import Blueprint, jsonify, request, session, send_file  # pyright: ignore[reportMissingImports]
from src.models import db, User, Curso, Inscripcion, LogActividad, CuposConfig, MunicipioCupo, Notificacion
from datetime import datetime, timedelta
import csv
import io
import tempfile
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

admin_bp = Blueprint('admin', __name__)

def require_admin(f):
    """Decorador para requerir rol de administrador"""
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return jsonify({'error': 'No autorizado'}), 401
        
        user = User.query.get(session['user_id'])
        if not user or user.rol != 'admin':
            return jsonify({'error': 'Acceso denegado. Se requiere rol de administrador'}), 403
        
        return f(*args, **kwargs)
    decorated_function.__name__ = f.__name__
    return decorated_function

@admin_bp.route('/dashboard/metrics', methods=['GET'])
@require_admin
def get_dashboard_metrics():
    """Obtener métricas generales del dashboard"""
    try:
        # Total de usuarios registrados
        total_usuarios = User.query.count()
        
        # Usuarios por estado
        usuarios_activos = User.query.filter_by(estado_cuenta='activa').count()
        usuarios_inactivos = User.query.filter_by(estado_cuenta='inactiva').count()
        usuarios_suspendidos = User.query.filter_by(estado_cuenta='suspendida').count()
        
        # Usuarios por rol
        estudiantes = User.query.filter_by(rol='estudiante').count()
        instructores = User.query.filter_by(rol='instructor').count()
        admins = User.query.filter_by(rol='admin').count()
        
        # Cursos activos
        cursos_activos = Curso.query.filter_by(activo=True).count()
        total_cursos = Curso.query.count()
        
        # Inscripciones (simulado por ahora)
        total_inscripciones = 0
        inscripciones_activas = 0
        
        # Usuarios registrados en los últimos 30 días
        fecha_limite = datetime.utcnow() - timedelta(days=30)
        nuevos_usuarios = User.query.filter(User.fecha_creacion >= fecha_limite).count()
        
        # Actividad reciente (simulado por ahora)
        actividad_reciente = 0
        
        return jsonify({
            'usuarios': {
                'total': total_usuarios,
                'activos': usuarios_activos,
                'inactivos': usuarios_inactivos,
                'suspendidos': usuarios_suspendidos,
                'nuevos_30_dias': nuevos_usuarios
            },
            'roles': {
                'estudiantes': estudiantes,
                'instructores': instructores,
                'administradores': admins
            },
            'cursos': {
                'total': total_cursos,
                'activos': cursos_activos
            },
            'inscripciones': {
                'total': total_inscripciones,
                'activas': inscripciones_activas
            },
            'actividad': {
                'ultimos_7_dias': actividad_reciente
            }
        }), 200
        
    except Exception as e:
        return jsonify({'error': f'Error al obtener métricas: {str(e)}'}), 500


@admin_bp.route('/cupos/config', methods=['GET', 'PUT'])
@require_admin
def cupos_config():
    try:
        if request.method == 'GET':
            cfg = CuposConfig.query.order_by(CuposConfig.id.desc()).first()
            return jsonify(cfg.to_dict() if cfg else {'modo': 'abierto', 'cupo_global_max': None, 'convocatoria': '2025'}), 200

        data = request.json or {}
        modo = (data.get('modo') or 'abierto').strip()
        if modo not in ['abierto', 'bloqueado']:
            return jsonify({'error': 'Modo inválido'}), 400
        cupo_global_max = data.get('cupo_global_max')
        if cupo_global_max is not None:
            try:
                cupo_global_max = int(cupo_global_max)
                if cupo_global_max < 0:
                    return jsonify({'error': 'cupo_global_max debe ser >= 0'}), 400
            except Exception:
                return jsonify({'error': 'cupo_global_max debe ser entero'}), 400

        convocatoria = (data.get('convocatoria') or '2025').strip()

        cfg = CuposConfig(modo=modo, cupo_global_max=cupo_global_max, convocatoria=convocatoria)
        db.session.add(cfg)
        db.session.commit()
        try:
            db.session.add(LogActividad(
                usuario_id=session.get('user_id'),
                accion='config_cupo_actualizada',
                detalles=f"modo={modo}, cupo_global_max={cupo_global_max}, convocatoria={convocatoria}"
            ))
            db.session.commit()
        except Exception:
            db.session.rollback()
        return jsonify({'message': 'Configuración actualizada', 'config': cfg.to_dict()}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Error al actualizar configuración: {str(e)}'}), 500


@admin_bp.route('/cupos/municipios', methods=['GET', 'PUT'])
@require_admin
def cupos_municipios():
    try:
        if request.method == 'GET':
            filas = MunicipioCupo.query.order_by(MunicipioCupo.subregion.asc(), MunicipioCupo.municipio_slug.asc()).all()
            return jsonify([f.to_dict() for f in filas]), 200

        data = request.json or {}
        items = data.get('items') or []
        with db.session.begin_nested():
            for it in items:
                slug = it.get('municipio_slug')
                cupo = it.get('cupo_max')
                if not slug:
                    continue
                try:
                    cupo = int(cupo)
                    if cupo < 0:
                        continue
                except Exception:
                    continue
                row = MunicipioCupo.query.filter_by(municipio_slug=slug).first()
                if row:
                    row.cupo_max = cupo
                else:
                    db.session.add(MunicipioCupo(municipio_slug=slug, subregion=it.get('subregion') or '', cupo_max=cupo))
        db.session.commit()
        try:
            db.session.add(LogActividad(
                usuario_id=session.get('user_id'),
                accion='municipio_cupo_actualizado',
                detalles=f"items={len(items)}"
            ))
            db.session.commit()
        except Exception:
            db.session.rollback()
        return jsonify({'message': 'Cupos de municipios actualizados'}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Error al actualizar cupos: {str(e)}'}), 500


@admin_bp.route('/cupos/estado', methods=['GET'])
@require_admin
def cupos_estado():
    """Estado de cupos: global, por subregión y por municipio"""
    try:
        # Configuración vigente
        cfg = CuposConfig.query.order_by(CuposConfig.id.desc()).first()
        convocatoria = cfg.convocatoria if cfg else '2025'

        # Cupos por municipio
        municipios = MunicipioCupo.query.all()
        muni_info = {m.municipio_slug: {'subregion': m.subregion, 'cupo_max': int(m.cupo_max)} for m in municipios}

        # Conteos por municipio
        q_base = db.session.query(User.municipio, User.estado_cuenta, db.func.count(User.id)).filter(
            User.convocatoria == convocatoria
        ).group_by(User.municipio, User.estado_cuenta).all()

        confirmados_map = {}
        espera_map = {}
        for muni, estado, cnt in q_base:
            if muni is None:
                continue
            if estado in ('activa', 'inactiva'):
                confirmados_map[muni] = confirmados_map.get(muni, 0) + int(cnt)
            elif estado == 'lista_espera':
                espera_map[muni] = espera_map.get(muni, 0) + int(cnt)

        # Construir lista municipal
        municipios_estado = []
        subregion_agg = {}
        total_cupo_municipal = 0
        total_confirmados = 0
        total_espera = 0
        for muni, info in muni_info.items():
            cupo_max = info['cupo_max']
            confirmados = confirmados_map.get(muni, 0)
            espera = espera_map.get(muni, 0)
            disponibles = max(0, cupo_max - confirmados)
            pct = (confirmados / cupo_max * 100.0) if cupo_max > 0 else 0.0
            municipios_estado.append({
                'municipio': muni,
                'subregion': info['subregion'],
                'cupo_max': cupo_max,
                'confirmados': confirmados,
                'lista_espera': espera,
                'disponibles': disponibles,
                'porcentaje': round(pct, 2),
            })

            sr = info['subregion']
            if sr not in subregion_agg:
                subregion_agg[sr] = {'cupo_max': 0, 'confirmados': 0, 'lista_espera': 0}
            subregion_agg[sr]['cupo_max'] += cupo_max
            subregion_agg[sr]['confirmados'] += confirmados
            subregion_agg[sr]['lista_espera'] += espera

            total_cupo_municipal += cupo_max
            total_confirmados += confirmados
            total_espera += espera

        subregiones_estado = []
        for sr, agg in subregion_agg.items():
            disp = max(0, agg['cupo_max'] - agg['confirmados'])
            pct = (agg['confirmados'] / agg['cupo_max'] * 100.0) if agg['cupo_max'] > 0 else 0.0
            subregiones_estado.append({
                'subregion': sr,
                'cupo_max': agg['cupo_max'],
                'confirmados': agg['confirmados'],
                'lista_espera': agg['lista_espera'],
                'disponibles': disp,
                'porcentaje': round(pct, 2),
            })

        # Global
        global_cupo = cfg.cupo_global_max if (cfg and cfg.cupo_global_max is not None) else total_cupo_municipal
        global_disponibles = max(0, global_cupo - total_confirmados)
        global_pct = (total_confirmados / global_cupo * 100.0) if global_cupo > 0 else 0.0

        return jsonify({
            'convocatoria': convocatoria,
            'global': {
                'cupo_max': int(global_cupo),
                'confirmados': int(total_confirmados),
                'lista_espera': int(total_espera),
                'disponibles': int(global_disponibles),
                'porcentaje': round(global_pct, 2),
            },
            'subregiones': subregiones_estado,
            'municipios': sorted(municipios_estado, key=lambda x: (x['subregion'], x['municipio']))
        }), 200
    except Exception as e:
        return jsonify({'error': f'Error al obtener estado de cupos: {str(e)}'}), 500


@admin_bp.route('/cupos/estado/export', methods=['GET'])
@require_admin
def cupos_estado_export():
    """Exportar estado de cupos en Excel"""
    try:
        # Reusar lógica del endpoint anterior
        cfg = CuposConfig.query.order_by(CuposConfig.id.desc()).first()
        convocatoria = cfg.convocatoria if cfg else '2025'
        municipios = MunicipioCupo.query.all()
        muni_info = {m.municipio_slug: {'subregion': m.subregion, 'cupo_max': int(m.cupo_max)} for m in municipios}
        q_base = db.session.query(User.municipio, User.estado_cuenta, db.func.count(User.id)).filter(
            User.convocatoria == convocatoria
        ).group_by(User.municipio, User.estado_cuenta).all()
        confirmados_map, espera_map = {}, {}
        for muni, estado, cnt in q_base:
            if muni is None:
                continue
            if estado in ('activa', 'inactiva'):
                confirmados_map[muni] = confirmados_map.get(muni, 0) + int(cnt)
            elif estado == 'lista_espera':
                espera_map[muni] = espera_map.get(muni, 0) + int(cnt)
        
        # Crear workbook de Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Estado de Cupos"
        
        # Estilos para el encabezado
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # Estilos para datos
        data_alignment = Alignment(horizontal="center", vertical="center")
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Verde para disponible
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")    # Rojo para agotado
        yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid") # Amarillo para ocupación alta
        
        # Encabezados
        headers = ['Subregión', 'Municipio', 'Cupo Máximo', 'Confirmados', 'Lista Espera', 'Disponibles', '% Ocupación']
        
        # Escribir encabezados con formato
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Datos con formato condicional
        for row, (muni, info) in enumerate(sorted(muni_info.items(), key=lambda x: (x[1]['subregion'], x[0])), 2):
            cupo_max = info['cupo_max']
            conf = confirmados_map.get(muni, 0)
            esp = espera_map.get(muni, 0)
            disp = max(0, cupo_max - conf)
            pct = round((conf / cupo_max * 100.0), 2) if cupo_max > 0 else 0.0
            
            data = [info['subregion'], muni, cupo_max, conf, esp, disp, f"{pct}%"]
            
            # Escribir datos con formato condicional
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.alignment = data_alignment
                cell.border = thin_border
                
                # Aplicar colores según ocupación (columna % Ocupación)
                if col == 7:  # Columna de porcentaje
                    if pct >= 100:
                        cell.fill = red_fill  # Rojo si está al 100% o más
                    elif pct >= 80:
                        cell.fill = yellow_fill  # Amarillo si está entre 80-99%
                    else:
                        cell.fill = green_fill  # Verde si está por debajo del 80%
        
        # Ajustar ancho de columnas
        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            ws.column_dimensions[column_letter].width = 15
        
        # Crear archivo en memoria
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.getvalue(), 200, {
            'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'Content-Disposition': 'attachment; filename=estado_cupos.xlsx'
        }
    except Exception as e:
        return jsonify({'error': f'Error al exportar: {str(e)}'}), 500

@admin_bp.route('/users', methods=['GET'])
@require_admin
def get_users():
    """Obtener lista de usuarios con filtros"""
    try:
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 20, type=int)
        estado = request.args.get('estado')
        rol = request.args.get('rol')
        search = request.args.get('search')
        estado_control = request.args.get('estado_control')
        
        query = User.query
        
        # Aplicar filtros
        if estado:
            query = query.filter_by(estado_cuenta=estado)
        if rol:
            query = query.filter_by(rol=rol)
        if estado_control:
            query = query.filter_by(estado_control=estado_control)
        if search:
            search_term = f"%{search}%"
            query = query.filter(
                (User.nombre.ilike(search_term)) |
                (User.apellido.ilike(search_term)) |
                (User.email.ilike(search_term)) |
                (User.numero_documento.ilike(search_term))
            )
        
        # Ordenar por fecha de creación (más recientes primero)
        query = query.order_by(User.fecha_creacion.desc())
        
        # Paginación
        pagination = query.paginate(
            page=page, 
            per_page=per_page, 
            error_out=False
        )
        
        users = pagination.items
        
        return jsonify({
            'users': [user.to_dict() for user in users],
            'pagination': {
                'page': page,
                'per_page': per_page,
                'total': pagination.total,
                'pages': pagination.pages,
                'has_next': pagination.has_next,
                'has_prev': pagination.has_prev
            }
        }), 200
        
    except Exception as e:
        return jsonify({'error': f'Error al obtener usuarios: {str(e)}'}), 500

@admin_bp.route('/users/<int:user_id>', methods=['PUT'])
@require_admin
def update_user(user_id):
    """Actualizar usuario (estado, rol, etc.)"""
    try:
        user = User.query.get_or_404(user_id)
        data = request.json
        
        # Campos permitidos para actualización
        if 'estado_cuenta' in data:
            user.estado_cuenta = data['estado_cuenta']
        if 'rol' in data:
            user.rol = data['rol']
        if 'nombre' in data:
            user.nombre = data['nombre']
        if 'apellido' in data:
            user.apellido = data['apellido']
        
        user.fecha_actualizacion = datetime.utcnow()
        
        # Registrar actividad (temporalmente comentado por problemas de mapeo)
        # log = LogActividad(
        #     usuario_id=session['user_id'],
        #     accion='actualizar_usuario',
        #     detalles=f'Usuario {user.email} actualizado por administrador'
        # )
        # db.session.add(log)
        
        db.session.commit()
        
        return jsonify({
            'message': 'Usuario actualizado exitosamente',
            'user': user.to_dict()
        }), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Error al actualizar usuario: {str(e)}'}), 500

@admin_bp.route('/users/bulk-update', methods=['POST'])
@require_admin
def bulk_update_users():
    """Actualización masiva de usuarios"""
    try:
        data = request.json
        user_ids = data.get('user_ids', [])
        updates = data.get('updates', {})
        
        if not user_ids:
            return jsonify({'error': 'No se proporcionaron IDs de usuarios'}), 400
        
        users = User.query.filter(User.id.in_(user_ids)).all()
        
        for user in users:
            if 'estado_cuenta' in updates:
                user.estado_cuenta = updates['estado_cuenta']
            if 'rol' in updates:
                user.rol = updates['rol']
            
            user.fecha_actualizacion = datetime.utcnow()
        
        # Registrar actividad (temporalmente comentado por problemas de mapeo)
        # log = LogActividad(
        #     usuario_id=session['user_id'],
        #     accion='actualizacion_masiva_usuarios',
        #     detalles=f'Actualización masiva de {len(users)} usuarios'
        # )
        # db.session.add(log)
        
        db.session.commit()
        
        return jsonify({
            'message': f'{len(users)} usuarios actualizados exitosamente'
        }), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Error en actualización masiva: {str(e)}'}), 500

@admin_bp.route('/users/import', methods=['POST'])
@require_admin
def import_users():
    """Importar usuarios desde CSV"""
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No se proporcionó archivo'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No se seleccionó archivo'}), 400
        
        if not file.filename.endswith('.csv'):
            return jsonify({'error': 'El archivo debe ser CSV'}), 400
        
        # Leer CSV
        stream = io.StringIO(file.stream.read().decode("UTF8"), newline=None)
        csv_reader = csv.DictReader(stream)
        
        imported_count = 0
        errors = []
        
        for row_num, row in enumerate(csv_reader, 2):  # Empezar en 2 porque la fila 1 es header
            try:
                # Validar campos requeridos
                required_fields = ['nombre', 'apellido', 'email', 'password', 'tipo_documento', 'numero_documento']
                for field in required_fields:
                    if not row.get(field):
                        errors.append(f"Fila {row_num}: Campo '{field}' es requerido")
                        continue
                
                # Verificar si el email ya existe
                if User.query.filter_by(email=row['email']).first():
                    errors.append(f"Fila {row_num}: Email '{row['email']}' ya existe")
                    continue
                
                # Crear usuario
                user = User(
                    nombre=row['nombre'],
                    apellido=row['apellido'],
                    email=row['email'],
                    tipo_documento=row['tipo_documento'],
                    numero_documento=row['numero_documento'],
                    rol=row.get('rol', 'estudiante'),
                    estado_cuenta=row.get('estado_cuenta', 'inactiva')
                )
                user.set_password(row['password'])
                
                db.session.add(user)
                imported_count += 1
                
            except Exception as e:
                errors.append(f"Fila {row_num}: Error - {str(e)}")
        
        if imported_count > 0:
            # Registrar actividad (temporalmente comentado por problemas de mapeo)
            # log = LogActividad(
            #     usuario_id=session['user_id'],
            #     accion='importar_usuarios',
            #     detalles=f'Importados {imported_count} usuarios desde CSV'
            # )
            # db.session.add(log)
            
            db.session.commit()
        
        return jsonify({
            'message': f'Importación completada',
            'imported_count': imported_count,
            'errors': errors
        }), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Error en importación: {str(e)}'}), 500

@admin_bp.route('/users/export', methods=['GET'])
@require_admin
def export_users():
    """Exportar todos los usuarios a Excel con formato profesional para auditorías"""
    try:
        # Obtener todos los usuarios
        users = User.query.all()
        
        # Crear workbook de Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Usuarios Registrados"
        
        # Crear segunda hoja para detalle de cursos
        ws_cursos = wb.create_sheet("Detalle Cursos por Usuario")
        
        # Estilos para el encabezado
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # Estilos para datos
        data_alignment = Alignment(horizontal="center", vertical="center")
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Verde claro para ✔️
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")    # Rojo claro para ❌
        yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid") # Amarillo para pendiente
        
        # Encabezados organizados según el formulario multistep + datos de cursos
        headers = [
            # PASO 1: Información básica
            'ID', 'Nombre', 'Apellido', 'Email', 'Municipio', 'Tipo Documento', 'Número Documento', 
            'Tipo Persona', 'Emprendimiento', 'Sector', 'Estado Cuenta', 'Fecha Creación',
            
            # PASO 2: Documentación obligatoria
            'TDR', 'Uso Imagen', 'Plan Negocio', 'Vecindad', 'Video',
            
            # PASO 3: Documentación según tipo de persona
            'RUT', 'Cédula', 'Cédula Representante', 'Cert. Existencia',
            
            # PASO 4: Documentación diferencial (subsanable)
            'RUV', 'SISBEN', 'Grupo Étnico', 'ARN', 'Discapacidad',
            
            # PASO 5: Documentación de control (obligatoria)
            'Fiscales', 'Disciplinarios', 'Judiciales', 'REDAM', 'Inhab. Sexuales', 'Capacidad Legal',
            'Estado Control', 'Resultado Certificados',
            
            # PASO 6: Certificación de funcionamiento
            'Formalizado', 'Matrícula Mercantil', 'Facturas 6M', 'Publicaciones Redes', 'Registro Ventas',
            
            # PASO 7: Financiación de otras fuentes
            'Financiado Estado', 'Regalías', 'Cámara Comercio', 'Incubadoras', 'Otro Financ.',
            
            # PASO 8: Declaraciones y aceptaciones
            'Declara Veraz', 'Declara No Beneficiario', 'Acepta Términos',
            
            # Estado del proceso
            'Paso Actual', 'Formulario Enviado', 'Estado Inscripción',
            
                # NUEVOS: Datos de cursos
                'Cursos Asignados', 'Cursos Completados', 'Avance Cursos (%)', 'Fase Actual',
                
                # NUEVOS: Datos de evidencias
                'Evidencias Estado', 'Tipo Emprendimiento'
        ]
        
        # Escribir encabezados con formato
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Datos de usuarios
        for row, user in enumerate(users, 2):
            # Función helper para estado de documento
            def doc_status(doc_field, is_optional=False):
                if doc_field:
                    return "✔️"
                elif is_optional:
                    return "⚠️"  # Pendiente/subsanable
                else:
                    return "❌"
            
            # Obtener datos de cursos del usuario
            from src.models import UsuarioCurso, EvidenciaFuncionamiento
            cursos_asignados = UsuarioCurso.query.filter_by(user_id=user.id).count()
            cursos_completados = UsuarioCurso.query.filter_by(user_id=user.id, estado='completado').count()
            avance_cursos = round((cursos_completados / cursos_asignados * 100) if cursos_asignados > 0 else 0, 2)
            
            # Datos de evidencias
            evidencia = EvidenciaFuncionamiento.query.filter_by(user_id=user.id).first()
            evidencias_estado = evidencia.estado_revision if evidencia else 'Sin evidencias'
            tipo_emprendimiento = evidencia.tipo if evidencia else 'N/A'
            
            # Datos básicos
            data = [
                user.id,
                user.nombre or '',
                user.apellido or '',
                user.email or '',
                user.municipio or '',
                user.tipo_documento or '',
                user.numero_documento or '',
                user.tipo_persona or '',
                user.emprendimiento_nombre or '',
                user.emprendimiento_sector or '',
                user.estado_cuenta or '',
                user.fecha_creacion.strftime('%Y-%m-%d %H:%M') if user.fecha_creacion else '',
                
                # Documentación obligatoria
                doc_status(user.doc_terminos_pdf),
                doc_status(user.doc_uso_imagen_pdf),
                doc_status(user.doc_plan_negocio_xls),
                doc_status(user.doc_vecindad_pdf),
                doc_status(user.video_url),
                
                # Documentación según tipo de persona
                doc_status(user.rut_pdf),
                doc_status(user.cedula_pdf),
                doc_status(user.cedula_representante_pdf),
                doc_status(user.cert_existencia_pdf),
                
                # Documentación diferencial (subsanable)
                doc_status(user.ruv_pdf, is_optional=True),
                doc_status(user.sisben_pdf, is_optional=True),
                doc_status(user.grupo_etnico_pdf, is_optional=True),
                doc_status(user.arn_pdf, is_optional=True),
                doc_status(user.discapacidad_pdf, is_optional=True),
                
                # Documentación de control
                doc_status(user.antecedentes_fiscales_pdf),
                doc_status(user.antecedentes_disciplinarios_pdf),
                doc_status(user.antecedentes_judiciales_pdf),
                doc_status(user.redam_pdf),
                doc_status(user.inhabilidades_sexuales_pdf),
                doc_status(user.declaracion_capacidad_legal_pdf),
                user.estado_control or 'pendiente',
                'RECHAZADO AUTOMÁTICO' if user.resultado_certificados == 'inhabilidad_detectada' and user.estado_cuenta == 'rechazada' else (user.resultado_certificados or 'pendiente'),
                
                # Certificación de funcionamiento
                'Sí' if user.emprendimiento_formalizado else 'No' if user.emprendimiento_formalizado is not None else 'N/A',
                doc_status(user.matricula_mercantil_pdf),
                doc_status(user.facturas_6meses_pdf),
                doc_status(user.publicaciones_redes_pdf),
                doc_status(user.registro_ventas_pdf),
                
                # Financiación
                'Sí' if user.financiado_estado else 'No' if user.financiado_estado is not None else 'N/A',
                '✔️' if user.financiado_regalias else '❌',
                '✔️' if user.financiado_camara_comercio else '❌',
                '✔️' if user.financiado_incubadoras else '❌',
                '✔️' if user.financiado_otro else '❌',
                
                # Declaraciones
                '✔️' if user.declara_veraz else '❌',
                '✔️' if user.declara_no_beneficiario else '❌',
                '✔️' if user.acepta_terminos else '❌',
                
                # Estado del proceso
                user.paso_actual or 1,
                'Sí' if user.formulario_enviado else 'No',
                user.estado_inscripcion or 'borrador',
                
                # Datos de cursos
                cursos_asignados,
                cursos_completados,
                f"{avance_cursos}%",
                getattr(user, 'fase_actual', 'N/A') or 'N/A',
                
                # Datos de evidencias
                evidencias_estado,
                tipo_emprendimiento
            ]
            
            # Escribir datos con formato condicional
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.alignment = data_alignment
                cell.border = thin_border
                
                # Aplicar colores según el valor
                if value == "✔️":
                    cell.fill = green_fill
                elif value == "❌":
                    cell.fill = red_fill
                elif value == "⚠️":
                    cell.fill = yellow_fill
        
        # Ajustar ancho de columnas
        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            ws.column_dimensions[column_letter].width = 15
        
        # === SEGUNDA HOJA: Detalle de Cursos por Usuario ===
        
        # Encabezados para la hoja de cursos
        headers_cursos = [
            'Usuario ID', 'Nombre Usuario', 'Email Usuario', 'Curso ID', 'Título Curso', 
            'Tipo Curso', 'Estado Curso', 'Progreso (%)', 'Fecha Asignación', 'Fecha Inicio', 
            'Fecha Completado', 'Fecha Última Actividad'
        ]
        
        # Escribir encabezados de cursos
        for col, header in enumerate(headers_cursos, 1):
            cell = ws_cursos.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Obtener todas las asignaciones de cursos
        from src.models import UsuarioCurso
        asignaciones = UsuarioCurso.query.join(User).join(Curso).all()
        
        # Escribir datos de cursos
        for row, asignacion in enumerate(asignaciones, 2):
            curso_data = [
                asignacion.user_id,
                f"{asignacion.user.nombre or ''} {asignacion.user.apellido or ''}".strip(),
                asignacion.user.email or '',
                asignacion.curso_id,
                asignacion.curso.titulo or '',
                asignacion.curso.tipo or '',
                asignacion.estado or '',
                f"{asignacion.progreso}%",
                asignacion.fecha_asignacion.strftime('%Y-%m-%d %H:%M') if asignacion.fecha_asignacion else '',
                asignacion.fecha_inicio.strftime('%Y-%m-%d %H:%M') if asignacion.fecha_inicio else '',
                asignacion.fecha_completado.strftime('%Y-%m-%d %H:%M') if asignacion.fecha_completado else '',
                asignacion.fecha_ultima_actividad.strftime('%Y-%m-%d %H:%M') if asignacion.fecha_ultima_actividad else ''
            ]
            
            # Escribir datos con formato
            for col, value in enumerate(curso_data, 1):
                cell = ws_cursos.cell(row=row, column=col, value=value)
                cell.alignment = data_alignment
                cell.border = thin_border
                
                # Aplicar colores según el estado del curso
                if asignacion.estado == 'completado':
                    cell.fill = green_fill
                elif asignacion.estado == 'en_progreso':
                    cell.fill = yellow_fill
                elif asignacion.estado == 'pendiente':
                    cell.fill = red_fill
        
        # Ajustar ancho de columnas para la hoja de cursos
        for col in range(1, len(headers_cursos) + 1):
            column_letter = get_column_letter(col)
            ws_cursos.column_dimensions[column_letter].width = 18
        
        # Crear archivo en memoria
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Registrar actividad
        try:
            log = LogActividad(
                usuario_id=session['user_id'],
                accion='exportar_usuarios_excel',
                detalles=f'Exportados {len(users)} usuarios a Excel'
            )
            db.session.add(log)
            db.session.commit()
        except Exception as log_error:
            print(f"Error al registrar log: {log_error}")
        
        return output.getvalue(), 200, {
            'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'Content-Disposition': 'attachment; filename=usuarios_registrados.xlsx'
        }
        
    except Exception as e:
        return jsonify({'error': f'Error en exportación: {str(e)}'}), 500

@admin_bp.route('/users/<int:user_id>/certificados', methods=['PUT'])
@require_admin
def actualizar_resultado_certificados(user_id):
    """Actualizar el resultado de los certificados de control de un usuario"""
    try:
        data = request.json
        resultado = data.get('resultado_certificados')
        
        # Validar resultado
        if resultado not in ['pendiente', 'limpio', 'inhabilidad_detectada']:
            return jsonify({'error': 'Resultado inválido. Debe ser: pendiente, limpio o inhabilidad_detectada'}), 400
        
        user = User.query.get_or_404(user_id)
        resultado_anterior = user.resultado_certificados
        
        # Actualizar resultado
        user.resultado_certificados = resultado
        
        # REGLA AUTOMÁTICA DE RECHAZO
        if resultado == 'inhabilidad_detectada':
            estado_anterior = user.estado_cuenta
            user.estado_cuenta = 'rechazada'
            
            # Log de auditoría del rechazo automático
            detalle_rechazo = f"RECHAZO AUTOMÁTICO: Usuario {user.nombre} {user.apellido} (ID: {user.id}) rechazado automáticamente por inhabilidad detectada en certificados de control. Estado anterior: {estado_anterior}. Resultado anterior: {resultado_anterior}"
            
            db.session.add(LogActividad(
                usuario_id=session['user_id'],
                accion='rechazo_automatico_certificados',
                detalles=detalle_rechazo
            ))
            
            # Log también en el usuario afectado
            db.session.add(LogActividad(
                usuario_id=user.id,
                accion='cuenta_rechazada_automaticamente',
                detalles=f"Cuenta rechazada automáticamente por inhabilidad detectada en certificados de control. Administrador que marcó: {session.get('user_email', 'N/A')}"
            ))
            
        elif resultado == 'limpio' and user.estado_cuenta == 'rechazada':
            # Si se marca como limpio y estaba rechazada, revertir a inactiva para revisión manual
            user.estado_cuenta = 'inactiva'
            
            db.session.add(LogActividad(
                usuario_id=session['user_id'],
                accion='revertir_rechazo_certificados',
                detalles=f"Rechazo revertido para usuario {user.nombre} {user.apellido} (ID: {user.id}). Certificados marcados como limpios. Estado cambiado a inactiva para revisión."
            ))
        
        db.session.commit()
        
        return jsonify({
            'message': f'Resultado de certificados actualizado a: {resultado}',
            'usuario': user.to_dict(),
            'rechazo_automatico_aplicado': resultado == 'inhabilidad_detectada'
        }), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Error al actualizar certificados: {str(e)}'}), 500

@admin_bp.route('/certificados/estadisticas', methods=['GET'])
@require_admin
def estadisticas_certificados():
    """Obtener estadísticas de certificados de control y rechazos automáticos"""
    try:
        # Estadísticas generales
        total_usuarios = User.query.count()
        control_completo = User.query.filter_by(estado_control='completo').count()
        control_pendiente = User.query.filter_by(estado_control='pendiente').count()
        
        # Resultados de certificados
        resultado_pendiente = User.query.filter_by(resultado_certificados='pendiente').count()
        resultado_limpio = User.query.filter_by(resultado_certificados='limpio').count()
        resultado_inhabilidad = User.query.filter_by(resultado_certificados='inhabilidad_detectada').count()
        
        # Rechazos automáticos
        rechazados_automaticamente = User.query.filter_by(estado_cuenta='rechazada', resultado_certificados='inhabilidad_detectada').count()
        
        # Logs de rechazos automáticos recientes
        rechazos_recientes = LogActividad.query.filter_by(accion='rechazo_automatico_certificados').order_by(LogActividad.fecha.desc()).limit(10).all()
        
        return jsonify({
            'estadisticas_generales': {
                'total_usuarios': total_usuarios,
                'control_completo': control_completo,
                'control_pendiente': control_pendiente
            },
            'resultados_certificados': {
                'pendiente': resultado_pendiente,
                'limpio': resultado_limpio,
                'inhabilidad_detectada': resultado_inhabilidad
            },
            'rechazos_automaticos': {
                'total': rechazados_automaticamente,
                'porcentaje': round((rechazados_automaticamente / total_usuarios * 100), 2) if total_usuarios > 0 else 0
            },
            'rechazos_recientes': [log.to_dict() for log in rechazos_recientes]
        }), 200
        
    except Exception as e:
        return jsonify({'error': f'Error al obtener estadísticas: {str(e)}'}), 500

@admin_bp.route('/logs', methods=['GET'])
@require_admin
def get_logs():
    """Obtener logs de actividad"""
    try:
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 50, type=int)
        usuario_id = request.args.get('usuario_id', type=int)
        accion = request.args.get('accion')
        
        # Por ahora retornamos logs vacíos para evitar errores
        logs = []
        
        return jsonify({
            'logs': logs,
            'pagination': {
                'page': page,
                'per_page': per_page,
                'total': 0,
                'pages': 1,
                'has_next': False,
                'has_prev': False
            }
        }), 200
        
    except Exception as e:
        return jsonify({'error': f'Error al obtener logs: {str(e)}'}), 500 

@admin_bp.route('/courses', methods=['GET'])
@require_admin
def get_courses():
    """Endpoint temporal - devuelve lista vacía para evitar errores"""
    return jsonify({
        'courses': [],
        'pagination': {
            'page': 1,
            'pages': 1,
            'per_page': 20,
            'total': 0,
            'has_next': False,
            'has_prev': False
        }
    }), 200

@admin_bp.route('/courses', methods=['POST'])
@require_admin
def create_course():
    """Crear nuevo curso"""
    try:
        data = request.json
        
        # Validar campos obligatorios
        if not data.get('titulo'):
            return jsonify({'error': 'El título del curso es obligatorio'}), 400
        
        # Procesar instructor_id
        instructor_id = data.get('instructor_id')
        if instructor_id == '' or instructor_id is None:
            instructor_id = None
        else:
            try:
                instructor_id = int(instructor_id)
            except (ValueError, TypeError):
                return jsonify({'error': 'ID de instructor inválido'}), 400
        
        # Crear nuevo curso
        curso = Curso(
            titulo=data['titulo'],
            descripcion=data.get('descripcion', ''),
            instructor_id=instructor_id,
            estado=data.get('estado', 'activo'),
            fecha_apertura=datetime.fromisoformat(data['fecha_apertura']) if data.get('fecha_apertura') else None,
            fecha_cierre=datetime.fromisoformat(data['fecha_cierre']) if data.get('fecha_cierre') else None,
            duracion_horas=data.get('duracion_horas', 0),
            nivel=data.get('nivel', 'básico'),
            categoria=data.get('categoria'),
            imagen_url=data.get('imagen_url'),
            max_estudiantes=data.get('max_estudiantes', 0)
        )
        
        db.session.add(curso)
        db.session.commit()
        
        # Obtener información del curso creado con instructor
        course_dict = curso.to_dict()
        if curso.instructor_id:
            instructor = User.query.get(curso.instructor_id)
            if instructor:
                course_dict['instructor_nombre'] = f"{instructor.nombre} {instructor.apellido}"
            else:
                course_dict['instructor_nombre'] = 'Instructor no encontrado'
        else:
            course_dict['instructor_nombre'] = 'Sin asignar'
        
        return jsonify({
            'message': 'Curso creado exitosamente',
            'course': course_dict
        }), 201
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Error al crear curso: {str(e)}'}), 500

@admin_bp.route('/courses/<int:course_id>', methods=['GET'])
@require_admin
def get_course(course_id):
    """Obtener curso específico"""
    try:
        curso = Curso.query.get_or_404(course_id)
        course_dict = curso.to_dict()
        
        # Obtener información del instructor
        if curso.instructor_id:
            instructor = User.query.get(curso.instructor_id)
            if instructor:
                course_dict['instructor_nombre'] = f"{instructor.nombre} {instructor.apellido}"
            else:
                course_dict['instructor_nombre'] = 'Instructor no encontrado'
        else:
            course_dict['instructor_nombre'] = 'Sin asignar'
        
        # Obtener total de inscripciones
        total_inscripciones = Inscripcion.query.filter_by(curso_id=curso.id).count()
        course_dict['total_inscripciones'] = total_inscripciones
        
        return jsonify(course_dict), 200
        
    except Exception as e:
        return jsonify({'error': f'Error al obtener curso: {str(e)}'}), 500

@admin_bp.route('/courses/<int:course_id>', methods=['PUT'])
@require_admin
def update_course(course_id):
    """Actualizar curso"""
    try:
        curso = Curso.query.get_or_404(course_id)
        data = request.json
        
        # Actualizar campos
        if 'titulo' in data:
            curso.titulo = data['titulo']
        if 'descripcion' in data:
            curso.descripcion = data['descripcion']
        if 'instructor_id' in data:
            # Procesar instructor_id
            instructor_id = data['instructor_id']
            if instructor_id == '' or instructor_id is None:
                curso.instructor_id = None
            else:
                try:
                    curso.instructor_id = int(instructor_id)
                except (ValueError, TypeError):
                    return jsonify({'error': 'ID de instructor inválido'}), 400
        if 'estado' in data:
            curso.estado = data['estado']
        if 'fecha_apertura' in data:
            curso.fecha_apertura = datetime.fromisoformat(data['fecha_apertura']) if data['fecha_apertura'] else None
        if 'fecha_cierre' in data:
            curso.fecha_cierre = datetime.fromisoformat(data['fecha_cierre']) if data['fecha_cierre'] else None
        if 'duracion_horas' in data:
            curso.duracion_horas = data['duracion_horas']
        if 'nivel' in data:
            curso.nivel = data['nivel']
        if 'categoria' in data:
            curso.categoria = data['categoria']
        if 'imagen_url' in data:
            curso.imagen_url = data['imagen_url']
        if 'max_estudiantes' in data:
            curso.max_estudiantes = data['max_estudiantes']
        
        curso.fecha_actualizacion = datetime.utcnow()
        
        db.session.commit()
        
        # Obtener información actualizada del curso con instructor
        course_dict = curso.to_dict()
        if curso.instructor_id:
            instructor = User.query.get(curso.instructor_id)
            if instructor:
                course_dict['instructor_nombre'] = f"{instructor.nombre} {instructor.apellido}"
            else:
                course_dict['instructor_nombre'] = 'Instructor no encontrado'
        else:
            course_dict['instructor_nombre'] = 'Sin asignar'
        
        return jsonify({
            'message': 'Curso actualizado exitosamente',
            'course': course_dict
        }), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Error al actualizar curso: {str(e)}'}), 500

@admin_bp.route('/courses/<int:course_id>', methods=['DELETE'])
@require_admin
def delete_course(course_id):
    """Eliminar curso"""
    try:
        curso = Curso.query.get_or_404(course_id)
        
        # Verificar si hay inscripciones
        if curso.inscripciones.count() > 0:
            return jsonify({'error': 'No se puede eliminar un curso que tiene estudiantes inscritos'}), 400
        
        db.session.delete(curso)
        db.session.commit()
        
        return jsonify({'message': 'Curso eliminado exitosamente'}), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Error al eliminar curso: {str(e)}'}), 500

@admin_bp.route('/instructors', methods=['GET'])
@require_admin
def get_instructors():
    """Obtener lista de instructores"""
    try:
        instructores = User.query.filter_by(rol='instructor').all()
        return jsonify([{
            'id': user.id,
            'nombre': f"{user.nombre} {user.apellido}",
            'email': user.email,
            'estado_cuenta': user.estado_cuenta
        } for user in instructores]), 200
        
    except Exception as e:
        return jsonify({'error': f'Error al obtener instructores: {str(e)}'}), 500 

@admin_bp.route('/users/<int:user_id>/documento', methods=['GET'])
@require_admin
def download_user_documento(user_id):
    """Descargar/ver documento de identidad PDF del usuario"""
    try:
        user = User.query.get_or_404(user_id)
        if not user.documento_pdf:
            return jsonify({'error': 'El usuario no tiene documento PDF cargado'}), 404
        
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.pdf')
        tmp.write(user.documento_pdf)
        tmp.flush()
        tmp.seek(0)
        
        filename = user.documento_pdf_nombre or 'documento.pdf'
        return send_file(tmp.name, mimetype='application/pdf', as_attachment=False, download_name=filename)
    except Exception as e:
        return jsonify({'error': f'Error al descargar documento: {str(e)}'}), 500

@admin_bp.route('/users/<int:user_id>/requisitos', methods=['GET'])
@require_admin
def download_user_requisitos(user_id):
    """Descargar/ver requisitos PDF del usuario"""
    try:
        user = User.query.get_or_404(user_id)
        if not user.requisitos_pdf:
            return jsonify({'error': 'El usuario no tiene requisitos PDF cargados'}), 404
        
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.pdf')
        tmp.write(user.requisitos_pdf)
        tmp.flush()
        tmp.seek(0)
        
        filename = user.requisitos_pdf_nombre or 'requisitos.pdf'
        return send_file(tmp.name, mimetype='application/pdf', as_attachment=False, download_name=filename)
    except Exception as e:
        return jsonify({'error': f'Error al descargar requisitos: {str(e)}'}), 500

# ===== ENDPOINTS PARA GESTIÓN DE FASES =====

@admin_bp.route('/stats/phases', methods=['GET'])
@require_admin
def get_phase_stats():
    """Obtener estadísticas de fases de usuarios"""
    try:
        # Estadísticas generales
        total_users = User.query.count()
        
        # Usuarios por fase
        inscripcion = User.query.filter_by(fase_actual='inscripcion').count()
        formacion = User.query.filter_by(fase_actual='formacion').count()
        entrega_activos = User.query.filter_by(fase_actual='entrega_activos').count()
        
        # Usuarios con fases completadas
        fases_completadas = User.query.filter_by(fase_completada=True).count()
        
        # Progreso por fase
        inscripcion_completada = User.query.filter_by(fase_actual='inscripcion', fase_completada=True).count()
        formacion_completada = User.query.filter_by(fase_actual='formacion', fase_completada=True).count()
        entrega_completada = User.query.filter_by(fase_actual='entrega_activos', fase_completada=True).count()
        
        return jsonify({
            'total_users': total_users,
            'inscripcion': inscripcion,
            'formacion': formacion,
            'entrega_activos': entrega_activos,
            'fases_completadas': fases_completadas,
            'progreso': {
                'inscripcion': {
                    'total': inscripcion,
                    'completada': inscripcion_completada,
                    'porcentaje': round((inscripcion_completada / inscripcion * 100) if inscripcion > 0 else 0, 2)
                },
                'formacion': {
                    'total': formacion,
                    'completada': formacion_completada,
                    'porcentaje': round((formacion_completada / formacion * 100) if formacion > 0 else 0, 2)
                },
                'entrega_activos': {
                    'total': entrega_activos,
                    'completada': entrega_completada,
                    'porcentaje': round((entrega_completada / entrega_activos * 100) if entrega_activos > 0 else 0, 2)
                }
            }
        }), 200
        
    except Exception as e:
        return jsonify({'error': f'Error al obtener estadísticas de fases: {str(e)}'}), 500

@admin_bp.route('/users', methods=['GET'])
@require_admin
def get_all_users():
    """Obtener todos los usuarios con información de fases"""
    try:
        users = User.query.all()
        
        users_data = []
        for user in users:
            users_data.append({
                'id': user.id,
                'nombre': user.nombre,
                'apellido': user.apellido,
                'email': user.email,
                'telefono': user.telefono,
                'municipio': user.municipio,
                'emprendimiento_nombre': user.emprendimiento_nombre,
                'emprendimiento_sector': user.emprendimiento_sector,
                'tipo_persona': user.tipo_persona,
                'fase_actual': getattr(user, 'fase_actual', 'inscripcion'),
                'fecha_entrada_fase': user.fecha_entrada_fase.isoformat() if hasattr(user, 'fecha_entrada_fase') and user.fecha_entrada_fase else None,
                'fase_completada': getattr(user, 'fase_completada', False),
                'estado_inscripcion': user.estado_inscripcion,
                'formulario_enviado': user.formulario_enviado,
                'fecha_creacion': user.fecha_creacion.isoformat() if user.fecha_creacion else None,
                'estado_cuenta': user.estado_cuenta,
                'rol': user.rol
            })
        
        return jsonify({'users': users_data}), 200
        
    except Exception as e:
        return jsonify({'error': f'Error al obtener usuarios: {str(e)}'}), 500

@admin_bp.route('/users/<int:user_id>/phase', methods=['POST'])
@require_admin
def update_user_phase(user_id):
    """Actualizar la fase de un usuario específico"""
    try:
        user = User.query.get_or_404(user_id)
        data = request.json
        nueva_fase = data.get('nueva_fase')
        
        # Validar fase válida
        fases_validas = ['inscripcion', 'formacion', 'entrega_activos']
        if nueva_fase not in fases_validas:
            return jsonify({'error': 'Fase no válida'}), 400
        
        # Obtener fase actual
        fase_actual = getattr(user, 'fase_actual', 'inscripcion')
        
        # Actualizar fase
        user.fase_actual = nueva_fase
        user.fecha_entrada_fase = datetime.utcnow()
        user.fase_completada = False
        
        db.session.commit()
        
        # Crear notificación de cambio de fase
        mensajes = {
            'inscripcion': 'Tu emprendimiento ha pasado a la fase de Inscripción y Selección',
            'formacion': 'Tu emprendimiento ha pasado a la fase de Formación',
            'entrega_activos': 'Tu emprendimiento ha pasado a la fase de Entrega de Activos Productivos'
        }
        
        mensaje = mensajes.get(nueva_fase, f'Tu emprendimiento ha pasado a la fase: {nueva_fase}')
        
        notificacion = Notificacion(
            user_id=user.id,
            mensaje=mensaje,
            fase_nueva=nueva_fase
        )
        db.session.add(notificacion)
        
        # Log de actividad
        db.session.add(LogActividad(
            usuario_id=user.id,
            accion='cambio_fase_admin',
            detalles=f"Administrador cambió fase de {fase_actual} a {nueva_fase}"
        ))
        db.session.commit()
        
        return jsonify({
            'message': f'Fase actualizada exitosamente a {nueva_fase}',
            'fase_actual': nueva_fase,
            'fecha_entrada_fase': user.fecha_entrada_fase.isoformat()
        }), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Error al actualizar fase: {str(e)}'}), 500

@admin_bp.route('/export/phases', methods=['GET'])
@require_admin
def export_phase_report():
    """Exportar reporte de fases en Excel"""
    try:
        users = User.query.all()
        
        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte de Fases"
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Encabezados
        headers = [
            'ID', 'Nombre', 'Apellido', 'Email', 'Teléfono', 'Municipio',
            'Emprendimiento', 'Sector', 'Tipo Persona', 'Fase Actual',
            'Fecha Entrada Fase', 'Fase Completada', 'Estado Inscripción',
            'Formulario Enviado', 'Fecha Creación', 'Estado Cuenta'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Datos
        for row, user in enumerate(users, 2):
            ws.cell(row=row, column=1, value=user.id)
            ws.cell(row=row, column=2, value=user.nombre)
            ws.cell(row=row, column=3, value=user.apellido)
            ws.cell(row=row, column=4, value=user.email)
            ws.cell(row=row, column=5, value=user.telefono)
            ws.cell(row=row, column=6, value=user.municipio)
            ws.cell(row=row, column=7, value=user.emprendimiento_nombre)
            ws.cell(row=row, column=8, value=user.emprendimiento_sector)
            ws.cell(row=row, column=9, value=user.tipo_persona)
            ws.cell(row=row, column=10, value=getattr(user, 'fase_actual', 'inscripcion'))
            ws.cell(row=row, column=11, value=user.fecha_entrada_fase.strftime('%Y-%m-%d %H:%M:%S') if hasattr(user, 'fecha_entrada_fase') and user.fecha_entrada_fase else 'N/A')
            ws.cell(row=row, column=12, value='Sí' if getattr(user, 'fase_completada', False) else 'No')
            ws.cell(row=row, column=13, value=user.estado_inscripcion)
            ws.cell(row=row, column=14, value='Sí' if user.formulario_enviado else 'No')
            ws.cell(row=row, column=15, value=user.fecha_creacion.strftime('%Y-%m-%d %H:%M:%S') if user.fecha_creacion else 'N/A')
            ws.cell(row=row, column=16, value=user.estado_cuenta)
        
        # Ajustar ancho de columnas
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        # Guardar en archivo temporal
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        wb.save(tmp.name)
        tmp.flush()
        tmp.seek(0)
        
        return send_file(
            tmp.name,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'reporte_fases_{datetime.now().strftime("%Y%m%d")}.xlsx'
        )
        
    except Exception as e:
        return jsonify({'error': f'Error al exportar reporte: {str(e)}'}), 500 